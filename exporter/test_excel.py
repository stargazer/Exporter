from exporter.excel import _clean, export
import openpyxl
import unittest


class Test_Clean(unittest.TestCase):

    def setUp(self):
        # Data to export
        self.data = []
        self.fields = []

        for i in range(0, 10):
            self.data.append('%i' %i)
            self.fields.append('id%i' %i)

        self.test_dict = dict(zip(self.fields, self.data))
        self.test_list = self.data
        self.test_set = sorted(set(self.data))
        self.test_tuple = tuple(self.data)

    def test_clean_dict(self):
        value = _clean(self.test_dict)

    def test_clean_list(self):
        value = _clean(self.test_list)
        self.assertEqual(value, '0, 1, 2, 3, 4, 5, 6, 7, 8, 9')

    def test_clean_set(self):
        value = _clean(self.test_set)
        self.assertEqual(value, '0, 1, 2, 3, 4, 5, 6, 7, 8, 9')

    def test_clean_tuple(self):
        value = _clean(self.test_tuple)
        self.assertEqual(value, '0, 1, 2, 3, 4, 5, 6, 7, 8, 9')


class TestExport(unittest.TestCase):

    def setUp(self):
        # Data to export
        self.data = []
        self.fields = []

        for i in range(0, 10):
            self.data.append('%i' %i)
            self.fields.append('id%i' %i)

        self.dict_data = dict(zip(self.fields, self.data))

    def test_get_excel_file(self):
        f = export(self.dict_data, self.fields, 'Test')
        wb = openpyxl.reader.excel.load_workbook(f)
        try:
            ws = wb.get_sheet_by_name('Test')
        except KeyError:
            ws = None

        self.assertTrue(ws)

    def test_excel_file_content(self):
        f = export(self.dict_data, self.fields, 'Test')
        wb = openpyxl.reader.excel.load_workbook(f)
        ws = wb.get_active_sheet()

        result = []
        for value in ws.values:
            result.append(value)

        self.assertEqual(result[1], (0L, 1L, 2L, 3L, 4L, 5L, 6L, 7L, 8L, 9L))

    def test_excel_file_is_xlsx(self):
        f = export(self.dict_data, self.fields, 'Test')
        wb = openpyxl.reader.excel.load_workbook(f)

        self.assertEqual(wb.mime_type, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml')

if __name__ == '__main__':
    unittest.main()
